from fastapi import FastAPI, Depends, HTTPException, Request, UploadFile
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import select, func
from pathlib import Path
from shutil import copyfileobj
from uuid import uuid4
from typing import List

from database import engine, Base, get_db
from models import User, UserRole, AuditLog
from discussion import DiscussionMessage

from uploads import router as uploads_router
from notifications import router as notifications_router
from notifications import notify_all_users, create_notification

from schemas import UserCreate, UserLogin, UserResponse, Token

from auth import hash_password, verify_password, create_access_token, get_current_user, require_role

# Create all tables
Base.metadata.create_all(bind=engine)

app = FastAPI()

UPLOAD_ROOT = Path(__file__).resolve().parent / "uploads"
DISCUSSION_UPLOAD_DIR = UPLOAD_ROOT / "discussion"
ALLOWED_DISCUSSION_EXTENSIONS = {".pdf", ".docx", ".jpg", ".jpeg", ".png"}
UPLOAD_ROOT.mkdir(exist_ok=True)
DISCUSSION_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

app.mount("/uploads", StaticFiles(directory=UPLOAD_ROOT), name="uploads")
app.include_router(uploads_router)
app.include_router(notifications_router)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:5174"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
def read_root():
    return {"message": "Expert Decision Replay Platform backend is running"}


@app.post("/register", response_model=UserResponse, tags=["Authentication"])
def register(user: UserCreate, db: Session = Depends(get_db)):
    existing_user = db.execute(
        select(User).where(User.email == user.email)
    ).scalar_one_or_none()

    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")

    new_user = User(
        full_name=user.full_name,
        email=user.email,
        hashed_password=hash_password(user.password),
        role=UserRole.employee,
    )

    db.add(new_user)
    db.commit()
    db.refresh(new_user)

    return new_user


@app.post("/login", response_model=Token, tags=["Authentication"])
def login(credentials: UserLogin, db: Session = Depends(get_db)):
    user = db.execute(
        select(User).where(User.email == credentials.email)
    ).scalar_one_or_none()

    if not user or not verify_password(credentials.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Invalid email or password")

    access_token = create_access_token(data={"sub": user.email, "role": user.role.value})

    return {"access_token": access_token, "token_type": "bearer"}


@app.get("/me", response_model=UserResponse, tags=["Authentication"])
def read_current_user(current_user: User = Depends(get_current_user)):
    return current_user

from schemas import ChangePassword

@app.put("/me/change-password")
def change_password(
    payload: ChangePassword,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not verify_password(payload.current_password, current_user.hashed_password):
        raise HTTPException(status_code=400, detail="Current password is incorrect")

    if len(payload.new_password) < 8:
        raise HTTPException(status_code=400, detail="New password must be at least 8 characters")

    current_user.hashed_password = hash_password(payload.new_password)
    db.commit()

    return {"message": "Password changed successfully"}


@app.get("/users", response_model=List[UserResponse], tags=["User Management"])
def list_users(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.admin))
):
    users = db.execute(select(User)).scalars().all()
    return users

from schemas import RoleUpdate

@app.put("/users/{user_id}/role", response_model=UserResponse, tags=["User Management"])
def update_user_role(
    user_id: int,
    role_update: RoleUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.admin)),
):
    target_user = db.execute(
        select(User).where(User.id == user_id)
    ).scalar_one_or_none()

    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")

    if target_user.id == current_user.id:
        raise HTTPException(
            status_code=400,
            detail="You cannot change your own role"
        )

    target_user.role = role_update.role
    db.commit()
    db.refresh(target_user)

    return target_user

from schemas import (
    DecisionCreate, DecisionResponse, DecisionStatusUpdate, DecisionReportResponse, ApprovalReportResponse, AuditReportResponse, )
from models import ( Decision, DecisionStatus, Approval, ApprovalAction, )
from typing import List as ListType

def _with_creator_name(decision: Decision) -> Decision:
    decision.creator_name = decision.creator.full_name if decision.creator else None
    return decision

@app.post("/decisions", response_model=DecisionResponse, tags=["Decision Management"])
def create_decision(
    decision: DecisionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    new_decision = Decision(
        title=decision.title,
        problem_statement=decision.problem_statement,
        category=decision.category,
        created_by=current_user.id,
        attachment_url=decision.attachment_url,
    )
    db.add(new_decision)
    db.commit()
    db.refresh(new_decision)
    new_decision.creator_name = new_decision.creator.full_name

    notify_all_users(
        db,
        exclude_user_id=current_user.id,
        title="New decision submitted",
        message=f"{current_user.full_name} created \"{new_decision.title}\"",
        type="DECISION_CREATED",
        link=f"/decisions/{new_decision.id}",
    )

    return new_decision

@app.get("/decisions", response_model=ListType[DecisionResponse], tags=["Decision Management"])
def list_decisions(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    user_id: int | None = None,
):
    if user_id is not None and current_user.role != UserRole.admin and current_user.id != user_id:
        raise HTTPException(
            status_code=403,
            detail="You can only request your own decisions."
        )

    query = select(Decision)
    if user_id is not None:
        query = query.where(Decision.created_by == user_id)

    decisions = db.execute(query).scalars().all()
    for d in decisions:
        d.creator_name = d.creator.full_name if d.creator else None
    return decisions

from datetime import datetime, timedelta

@app.get("/decisions/escalations", response_model=ListType[DecisionResponse], tags=["Approval Workflow"])
def get_escalated_decisions(
    hours: int = 48,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.admin)),
):
    cutoff = datetime.utcnow() - timedelta(hours=hours)

    decisions = db.execute(
        select(Decision).where(
            Decision.status == DecisionStatus.under_review,
            Decision.created_at <= cutoff,
        )
    ).scalars().all()

    for d in decisions:
        d.creator_name = d.creator.full_name if d.creator else None

    return decisions

@app.get("/decisions/{decision_id}", response_model=DecisionResponse, tags=["Decision Management"])
def get_decision(
    decision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    decision = db.execute(
        select(Decision).where(Decision.id == decision_id)
    ).scalar_one_or_none()
    if not decision:
        raise HTTPException(status_code=404, detail="Decision not found")
    decision.creator_name = decision.creator.full_name if decision.creator else None
    return decision


@app.put("/decisions/{decision_id}/status", response_model=DecisionResponse, tags=["Decision Management"])
def update_decision_status(
    decision_id: int,
    status_update: DecisionStatusUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.manager, UserRole.admin)),
):
    decision = db.execute(
        select(Decision).where(Decision.id == decision_id)
    ).scalar_one_or_none()
    if not decision:
        raise HTTPException(status_code=404, detail="Decision not found")

    if status_update.status in (DecisionStatus.approved, DecisionStatus.rejected):
        raise HTTPException(
            status_code=400,
            detail="Use the /approve or /reject endpoints to move a decision to approved or rejected, so the review is properly recorded with a comment and reviewer."
        )

    decision.status = status_update.status
    db.commit()
    db.refresh(decision)
    return decision

from schemas import ApprovalCreate, ApprovalResponse
from models import Approval, ApprovalAction

@app.post("/decisions/{decision_id}/approve", response_model=ApprovalResponse, tags=["Approval Workflow"])
def approve_decision(
    decision_id: int,
    approval: ApprovalCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    decision = db.execute(
        select(Decision).where(Decision.id == decision_id)
    ).scalar_one_or_none()
    if not decision:
        raise HTTPException(status_code=404, detail="Decision not found")

    last_approval = db.execute(
        select(Approval)
        .where(Approval.decision_id == decision_id, Approval.action == ApprovalAction.approved)
        .order_by(Approval.stage.desc())
    ).scalars().first()

    next_stage = 1 if not last_approval else last_approval.stage + 1

    if next_stage == 1:
        if current_user.role not in (UserRole.reviewer, UserRole.manager, UserRole.admin):
            raise HTTPException(status_code=403, detail="Only a Reviewer, Manager, or Admin can give the initial approval")
    elif next_stage == 2:
        if current_user.role not in (UserRole.manager, UserRole.admin):
            raise HTTPException(status_code=403, detail="Only a Manager or Admin can give final approval")
    else:
        raise HTTPException(status_code=400, detail="This decision has already completed both approval stages")
    
    new_approval = Approval(
        decision_id=decision_id,
        reviewer_id=current_user.id,
        action=ApprovalAction.approved,
        comment=approval.comment,
        stage=next_stage,
    )
    db.add(new_approval)

    if next_stage == 2:
        decision.status = DecisionStatus.approved

    db.commit()
    db.refresh(new_approval)
    new_approval.reviewer_name = current_user.full_name

    if decision.status == DecisionStatus.approved:
        create_notification(
            db,
            user_id=decision.created_by,
            title="Decision approved",
            message=f"Your decision \"{decision.title}\" was approved by {current_user.full_name}",
            type="DECISION_APPROVED",
            link=f"/decisions/{decision.id}",
        )
    else:
        create_notification(
            db,
            user_id=decision.created_by,
            title="Decision moved to next approval stage",
            message=f"\"{decision.title}\" was approved by {current_user.full_name} and needs one more approval",
            type="DECISION_CREATED",
            link=f"/decisions/{decision.id}",
        )

    return new_approval


@app.post("/decisions/{decision_id}/reject", response_model=ApprovalResponse, tags=["Approval Workflow"])
def reject_decision(
    decision_id: int,
    approval: ApprovalCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.reviewer, UserRole.manager, UserRole.admin)),
):
    decision = db.execute(
        select(Decision).where(Decision.id == decision_id)
    ).scalar_one_or_none()
    if not decision:
        raise HTTPException(status_code=404, detail="Decision not found")

    if not approval.comment or not approval.comment.strip():
        raise HTTPException(status_code=400, detail="A comment is required when rejecting a decision")

    last_approval = db.execute(
        select(Approval)
        .where(Approval.decision_id == decision_id, Approval.action == ApprovalAction.approved)
        .order_by(Approval.stage.desc())
    ).scalars().first()
    current_stage = 1 if not last_approval else last_approval.stage + 1

    new_approval = Approval(
        decision_id=decision_id,
        reviewer_id=current_user.id,
        action=ApprovalAction.rejected,
        comment=approval.comment,
        stage=current_stage,
    )
    db.add(new_approval)

    decision.status = DecisionStatus.rejected
    db.commit()
    db.refresh(new_approval)
    new_approval.reviewer_name = current_user.full_name

    create_notification(
        db,
        user_id=decision.created_by,
        title="Decision rejected",
        message=f"Your decision \"{decision.title}\" was rejected by {current_user.full_name}: {approval.comment}",
        type="DECISION_REJECTED",
        link=f"/decisions/{decision.id}",
    )

    return new_approval

@app.post("/decisions/{decision_id}/resubmit", response_model=DecisionResponse, tags=["Approval Workflow"])
def resubmit_decision(
    decision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    decision = db.execute(
        select(Decision).where(Decision.id == decision_id)
    ).scalar_one_or_none()
    if not decision:
        raise HTTPException(status_code=404, detail="Decision not found")

    if decision.status != DecisionStatus.rejected:
        raise HTTPException(
            status_code=400,
            detail="Only a rejected decision can be resubmitted for review"
        )

    if current_user.id != decision.created_by and current_user.role != UserRole.admin:
        raise HTTPException(
            status_code=403,
            detail="Only the decision's creator or an admin can resubmit it"
        )

    last_rejection = db.execute(
        select(Approval)
        .where(Approval.decision_id == decision_id, Approval.action == ApprovalAction.rejected)
        .order_by(Approval.created_at.desc())
    ).scalars().first()

    if last_rejection:
        edited_after_rejection = db.execute(
            select(DecisionVersion)
            .where(DecisionVersion.decision_id == decision_id, DecisionVersion.created_at > last_rejection.created_at)
        ).scalars().first()

        if not edited_after_rejection:
            raise HTTPException(
                status_code=400,
                detail="Please edit the decision to address the rejection feedback before resubmitting."
            )
    
    resubmission_record = Approval(
        decision_id=decision_id,
        reviewer_id=current_user.id,
        action=ApprovalAction.resubmitted,
        comment="Resubmitted for review after rejection",
    )
    db.add(resubmission_record)

    decision.status = DecisionStatus.under_review
    db.commit()
    db.refresh(decision)

    notify_all_users(
        db,
        exclude_user_id=current_user.id,
        title="Decision resubmitted",
        message=f"\"{decision.title}\" was resubmitted for review by {current_user.full_name}",
        type="DECISION_CREATED",
        link=f"/decisions/{decision.id}",
    )

    return decision

@app.get("/decisions/{decision_id}/approvals", response_model=ListType[ApprovalResponse], tags=["Approval Workflow"])
def get_decision_approvals(
    decision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    decision = db.execute(
        select(Decision).where(Decision.id == decision_id)
    ).scalar_one_or_none()
    if not decision:
        raise HTTPException(status_code=404, detail="Decision not found")

    approvals = db.execute(
        select(Approval)
        .where(Approval.decision_id == decision_id)
        .order_by(Approval.created_at)
    ).scalars().all()
    for a in approvals:
        a.reviewer_name = a.reviewer.full_name if a.reviewer else None
    return approvals

from schemas import DecisionUpdate, DecisionVersionResponse
from models import DecisionVersion

@app.put("/decisions/{decision_id}", response_model=DecisionResponse, tags=["Decision Management"])
def update_decision(
    decision_id: int,
    decision_update: DecisionUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    decision = db.execute(
        select(Decision).where(Decision.id == decision_id)
    ).scalar_one_or_none()

    if not decision:
        raise HTTPException(status_code=404, detail="Decision not found")

    update_data = decision_update.model_dump(exclude_unset=True)
    if "attachment_url" in update_data and decision.attachment_url:
        if current_user.id != decision.created_by and current_user.role != UserRole.admin:
            raise HTTPException(
                status_code=403,
                detail="Only the decision's creator or an admin can change or remove its attachment"
            )

    existing_versions = db.execute(
        select(DecisionVersion).where(DecisionVersion.decision_id == decision_id)
    ).scalars().all()
    next_version_number = len(existing_versions) + 1

    snapshot = DecisionVersion(
        decision_id=decision.id,
        version_number=next_version_number,
        title=decision.title,
        problem_statement=decision.problem_statement,
        category=decision.category,
        status=decision.status.value,
        changed_by=current_user.id,
    )
    db.add(snapshot)

    for key, value in update_data.items():
        setattr(decision, key, value)

    db.commit()
    db.refresh(decision)
    return decision


@app.get("/decisions/{decision_id}/versions", response_model=ListType[DecisionVersionResponse], tags=["Version Tracking"])
def get_decision_versions(
    decision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    decision = db.execute(
        select(Decision).where(Decision.id == decision_id)
    ).scalar_one_or_none()

    if not decision:
        raise HTTPException(status_code=404, detail="Decision not found")

    versions = db.execute(
        select(DecisionVersion)
        .where(DecisionVersion.decision_id == decision_id)
        .order_by(DecisionVersion.version_number)
    ).scalars().all()

    return versions

@app.delete("/decisions/{decision_id}", tags=["Decision Management"])
def delete_decision(
    decision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.admin)),
):
    decision = db.execute(
        select(Decision).where(Decision.id == decision_id)
    ).scalar_one_or_none()

    if not decision:
        raise HTTPException(status_code=404, detail="Decision not found")

    db.delete(decision)
    db.commit()

    return {"message": "Decision deleted successfully"}

from fastapi.responses import StreamingResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
import io

@app.get("/decisions/{decision_id}/export", tags=["Decision Management"])
def export_decision_pdf(
    decision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    decision = db.execute(
        select(Decision).where(Decision.id == decision_id)
    ).scalar_one_or_none()
    if not decision:
        raise HTTPException(status_code=404, detail="Decision not found")

    alternatives = db.execute(
        select(Alternative).where(Alternative.decision_id == decision_id)
    ).scalars().all()

    creator_name = decision.creator.full_name if decision.creator else "Unknown"

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    y = height - 1 * inch

    def line(text, size=11, bold=False, gap=0.28):
        nonlocal y
        p.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        p.drawString(1 * inch, y, text)
        y -= gap * inch

    line("Expert Decision Replay Platform", size=10, bold=True)
    line("Decision Record", size=16, bold=True, gap=0.4)

    line(f"Title: {decision.title}", bold=True)
    line(f"Status: {decision.status.value.replace('_', ' ').title()}")
    line(f"Category: {decision.category or 'Uncategorized'}")
    line(f"Created by: {creator_name}")
    line(f"Created on: {decision.created_at.strftime('%d %b %Y, %I:%M %p')}", gap=0.4)

    line("Problem Statement:", bold=True)
    statement = decision.problem_statement
    words = statement.split()
    current_line = ""
    for word in words:
        if len(current_line) + len(word) < 90:
            current_line += word + " "
        else:
            line(current_line)
            current_line = word + " "
    if current_line:
        line(current_line, gap=0.4)

    if alternatives:
        line("Alternatives Considered:", bold=True)
        for alt in alternatives:
            line(f"- {alt.title}  (Risk: {alt.risk_level.value}, Feasibility: {alt.feasibility.value}, Cost: {alt.cost})")
        y -= 0.1 * inch

    p.setFont("Helvetica-Oblique", 8)
    p.drawString(1 * inch, 0.6 * inch, "Generated from Expert Decision Replay Platform")

    p.save()
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=decision_{decision_id}.pdf"}
    )


# ===================== ALTERNATIVES ENDPOINTS =====================

from schemas import AlternativeCreate, AlternativeResponse, AlternativeUpdate, AlternativeComparisonResponse
from models import Alternative

@app.post("/alternatives", response_model=AlternativeResponse, tags=["Alternative Comparison"])
def create_alternative(
    alternative: AlternativeCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    decision = db.execute(
        select(Decision).where(Decision.id == alternative.decision_id)
    ).scalar_one_or_none()

    if not decision:
        raise HTTPException(status_code=404, detail="Decision not found")

    new_alternative = Alternative(
        decision_id=alternative.decision_id,
        title=alternative.title,
        description=alternative.description,
        pros=alternative.pros,
        cons=alternative.cons,
        cost=alternative.cost,
        risk_level=alternative.risk_level,
        feasibility=alternative.feasibility,
    )

    db.add(new_alternative)
    db.commit()
    db.refresh(new_alternative)

    return new_alternative


@app.get("/decisions/{decision_id}/alternatives", response_model=ListType[AlternativeResponse], tags=["Alternative Comparison"])
def list_alternatives(
    decision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    decision = db.execute(
        select(Decision).where(Decision.id == decision_id)
    ).scalar_one_or_none()

    if not decision:
        raise HTTPException(status_code=404, detail="Decision not found")

    alternatives = db.execute(
        select(Alternative).where(Alternative.decision_id == decision_id)
    ).scalars().all()

    return alternatives


@app.get("/alternatives/{alternative_id}", response_model=AlternativeResponse, tags=["Alternative Comparison"])
def get_alternative(
    alternative_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    alternative = db.execute(
        select(Alternative).where(Alternative.id == alternative_id)
    ).scalar_one_or_none()

    if not alternative:
        raise HTTPException(status_code=404, detail="Alternative not found")

    return alternative


@app.put("/alternatives/{alternative_id}", response_model=AlternativeResponse, tags=["Alternative Comparison"])
def update_alternative(
    alternative_id: int,
    alternative_update: AlternativeUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    alternative = db.execute(
        select(Alternative).where(Alternative.id == alternative_id)
    ).scalar_one_or_none()
    if not alternative:
        raise HTTPException(status_code=404, detail="Alternative not found")

    update_data = alternative_update.model_dump(exclude_unset=True)

    for key, value in update_data.items():
        setattr(alternative, key, value)

    db.commit()
    db.refresh(alternative)

    return alternative


@app.delete("/alternatives/{alternative_id}", tags=["Alternative Comparison"])
def delete_alternative(
    alternative_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    alternative = db.execute(
        select(Alternative).where(Alternative.id == alternative_id)
    ).scalar_one_or_none()

    if not alternative:
        raise HTTPException(status_code=404, detail="Alternative not found")

    db.delete(alternative)
    db.commit()

    return {"message": "Alternative deleted successfully"}


@app.get("/decisions/{decision_id}/compare", response_model=AlternativeComparisonResponse, tags=["Alternative Comparison"])
def compare_alternatives(
    decision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    decision = db.execute(
        select(Decision).where(Decision.id == decision_id)
    ).scalar_one_or_none()

    if not decision:
        raise HTTPException(status_code=404, detail="Decision not found")

    alternatives = db.execute(
        select(Alternative).where(Alternative.decision_id == decision_id)
    ).scalars().all()

    return {
        "decision_id": decision.id,
        "decision_title": decision.title,
        "alternatives": alternatives,
    }


# ===================== DISCUSSION ENDPOINTS =====================

from schemas import DiscussionCreate, DiscussionReplyCreate, DiscussionUpdate, DiscussionResponse
from crud_discussion import (
    add_comment,
    add_meeting_note,
    delete_comment,
    edit_comment,
    get_comment_or_none,
    get_comments_for_decision,
    get_decision_or_none,
    reply_to_comment,
)


def _read_discussion_payload(payload: dict, required_fields: list[str]) -> dict:
    missing_fields = [field for field in required_fields if payload.get(field) in (None, "")]
    if missing_fields:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required field(s): {', '.join(missing_fields)}",
        )
    return payload


async def _parse_discussion_request(request: Request) -> tuple[dict, UploadFile | None]:
    content_type = request.headers.get("content-type", "")
    if content_type.startswith("multipart/form-data"):
        form = await request.form()
        attachment = form.get("attachment")
        payload = {key: value for key, value in form.items() if key != "attachment"}
        if attachment is not None and not hasattr(attachment, "filename"):
            attachment = None
        return payload, attachment

    if content_type.startswith("application/json"):
        return await request.json(), None

    raise HTTPException(status_code=415, detail="Use JSON or multipart/form-data")


def _save_discussion_attachment(attachment: UploadFile | None) -> str | None:
    if not attachment or not attachment.filename:
        return None

    extension = Path(attachment.filename).suffix.lower()
    if extension not in ALLOWED_DISCUSSION_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail="Attachment must be a PDF, DOCX, JPG, or PNG file",
        )

    filename = f"{uuid4().hex}{extension}"
    file_path = DISCUSSION_UPLOAD_DIR / filename
    with file_path.open("wb") as buffer:
        copyfileobj(attachment.file, buffer)

    return f"/uploads/discussion/{filename}"


@app.post(
    "/discussion",
    response_model=DiscussionResponse,
    tags=["Discussion Module"],
)
async def add_discussion_comment(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    payload, attachment = await _parse_discussion_request(request)
    payload = _read_discussion_payload(payload, ["decision_id", "message"])
    discussion = DiscussionCreate(
        decision_id=int(str(payload["decision_id"])),
        message=str(payload["message"]),
        attachment_url=payload.get("attachment_url"),
    )

    if not get_decision_or_none(db, discussion.decision_id):
        raise HTTPException(status_code=404, detail="Decision not found")

    attachment_url = _save_discussion_attachment(attachment) or discussion.attachment_url
    new_comment = add_comment(
        db,
        decision_id=discussion.decision_id,
        user_id=current_user.id,
        message=discussion.message,
        attachment_url=attachment_url,
    )

    decision_for_notif = get_decision_or_none(db, discussion.decision_id)
    notify_all_users(
        db,
        exclude_user_id=current_user.id,
        title="New discussion comment",
        message=f"{current_user.full_name} commented on \"{decision_for_notif.title}\"",
        type="NEW_DISCUSSION",
        link=f"/decisions/{discussion.decision_id}",
    )

    return new_comment


@app.post(
    "/discussion/meeting-note",
    response_model=DiscussionResponse,
    tags=["Discussion Module"],
)
async def add_discussion_meeting_note(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    payload, attachment = await _parse_discussion_request(request)
    payload = _read_discussion_payload(payload, ["decision_id", "message"])
    decision_id = int(str(payload["decision_id"]))
    message = str(payload["message"])
    attachment_url = payload.get("attachment_url")

    if not get_decision_or_none(db, decision_id):
        raise HTTPException(status_code=404, detail="Decision not found")

    saved_attachment_url = _save_discussion_attachment(attachment) or attachment_url
    return add_meeting_note(
        db,
        decision_id=decision_id,
        user_id=current_user.id,
        message=message,
        attachment_url=saved_attachment_url,
    )


@app.post(
    "/discussion/reply",
    response_model=DiscussionResponse,
    tags=["Discussion Module"],
)
async def reply_to_discussion_comment(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    payload, attachment = await _parse_discussion_request(request)
    payload = _read_discussion_payload(payload, ["parent_id", "message"])
    discussion_reply = DiscussionReplyCreate(
        parent_id=int(payload["parent_id"]),
        message=str(payload["message"]),
        attachment_url=payload.get("attachment_url"),
    )

    parent = get_comment_or_none(db, discussion_reply.parent_id)
    if not parent:
        raise HTTPException(status_code=404, detail="Parent comment not found")
    if not get_decision_or_none(db, parent.decision_id):
        raise HTTPException(status_code=404, detail="Decision not found")

    attachment_url = _save_discussion_attachment(attachment) or discussion_reply.attachment_url
    return reply_to_comment(
        db,
        parent=parent,
        user_id=current_user.id,
        message=discussion_reply.message,
        attachment_url=attachment_url,
    )


@app.get("/discussion/decision/{decision_id}", response_model=ListType[DiscussionResponse], tags=["Discussion Module"])
def get_decision_discussion_thread(
    decision_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not get_decision_or_none(db, decision_id):
        raise HTTPException(status_code=404, detail="Decision not found")
    return get_comments_for_decision(db, decision_id)


@app.put(
    "/discussion/{discussion_id}",
    response_model=DiscussionResponse, 
    tags=["Discussion Module"],
)
async def edit_discussion_comment(
    discussion_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    comment = get_comment_or_none(db, discussion_id)
    if not comment:
        raise HTTPException(status_code=404, detail="Comment not found")
    if comment.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Only the comment owner can edit this comment")

    payload, attachment = await _parse_discussion_request(request)
    update = DiscussionUpdate(
        message=payload.get("message"),
        attachment_url=payload.get("attachment_url"),
    )
    attachment_url = _save_discussion_attachment(attachment) or update.attachment_url
    if update.message is None and attachment_url is None:
        raise HTTPException(status_code=400, detail="Provide message or attachment")

    return edit_comment(
        db,
        comment=comment,
        user_id=current_user.id,
        message=update.message,
        attachment_url=attachment_url,
    )


@app.delete("/discussion/{discussion_id}", tags=["Discussion Module"])
def delete_discussion_comment(
    discussion_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    comment = get_comment_or_none(db, discussion_id)
    if not comment:
        raise HTTPException(status_code=404, detail="Comment not found")

    can_delete = comment.user_id == current_user.id or current_user.role in {
        UserRole.manager,
        UserRole.admin,
    }
    if not can_delete:
        raise HTTPException(status_code=403, detail="Only the owner, manager, or admin can delete this comment")

    delete_comment(db, comment=comment, user=current_user)
    return {"message": "Comment deleted successfully"}

# ============================= DASHBOARD ENDPOINTS =====================

from schemas import (
    EmployeeDashboardResponse,
    ReviewerDashboardResponse,
    ManagerDashboardResponse,
    AdminDashboardResponse,
)

def get_dashboard_stats(
    db: Session,
    user_id: int | None = None,
):
    filters = []

    if user_id is not None:
        filters.append(
            Decision.created_by == user_id
        )

    total_decisions = db.execute(
        select(func.count(Decision.id))
        .where(*filters)
    ).scalar_one()

    status_counts = dict(
        db.execute(
            select(
                Decision.status,
                func.count(Decision.id)
            )
            .where(*filters)
            .group_by(Decision.status)
        ).all()
    )

    recent_decisions = db.execute(
        select(Decision)
        .where(*filters)
        .order_by(Decision.created_at.desc())
        .limit(5)
    ).scalars().all()

    return {
        "total_decisions": total_decisions,
        "draft_decisions": status_counts.get(DecisionStatus.draft, 0),
        "under_review_decisions": status_counts.get(
            DecisionStatus.under_review, 0
        ),
        "approved_decisions": status_counts.get(
            DecisionStatus.approved, 0
        ),
        "rejected_decisions": status_counts.get(
            DecisionStatus.rejected, 0
        ),
        "archived_decisions": status_counts.get(
            DecisionStatus.archived, 0
        ),
        "recent_decisions": recent_decisions,
    }
    
@app.get(
    "/dashboard/employee",
    response_model=EmployeeDashboardResponse,
    tags=["Dashboard"],
)
def employee_dashboard(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    stats = get_dashboard_stats(
        db,
        current_user.id,
    )

    return EmployeeDashboardResponse(
        my_decisions=stats["total_decisions"],
        draft_decisions=stats["draft_decisions"],
        under_review_decisions=stats["under_review_decisions"],
        approved_decisions=stats["approved_decisions"],
        rejected_decisions=stats["rejected_decisions"],
        archived_decisions=stats["archived_decisions"],
        recent_decisions=stats["recent_decisions"],
    )
    
@app.get(
    "/dashboard/reviewer",
    response_model=ReviewerDashboardResponse,
    tags=["Dashboard"],
)
def reviewer_dashboard(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.reviewer:
        raise HTTPException(
            status_code=403,
            detail="Only reviewers can access this dashboard."
        )

    under_review_decisions = db.execute(
        select(func.count(Decision.id))
        .where(
            Decision.status == DecisionStatus.under_review
        )
    ).scalar_one()

    approved_decisions = db.execute(
        select(func.count(Decision.id))
        .where(
            Decision.status == DecisionStatus.approved
        )
    ).scalar_one()

    rejected_decisions = db.execute(
        select(func.count(Decision.id))
        .where(
            Decision.status == DecisionStatus.rejected
        )
    ).scalar_one()

    recent_under_review = db.execute(
        select(Decision)
        .where(
            Decision.status == DecisionStatus.under_review
        )
        .order_by(Decision.created_at.desc())
        .limit(5)
    ).scalars().all()

    return ReviewerDashboardResponse(
        under_review_decisions=under_review_decisions,
        approved_decisions=approved_decisions,
        rejected_decisions=rejected_decisions,
        recent_under_review=recent_under_review,
    )
    
@app.get(
    "/dashboard/manager",
    response_model=ManagerDashboardResponse,
    tags=["Dashboard"],
)
def manager_dashboard(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.manager:
        raise HTTPException(
            status_code=403,
            detail="Only managers can access this dashboard."
        )

    stats = get_dashboard_stats(db)

    return ManagerDashboardResponse(
        total_decisions=stats["total_decisions"],
        draft_decisions=stats["draft_decisions"],
        under_review_decisions=stats["under_review_decisions"],
        approved_decisions=stats["approved_decisions"],
        rejected_decisions=stats["rejected_decisions"],
        archived_decisions=stats["archived_decisions"],
        recent_decisions=stats["recent_decisions"],
    )
    
@app.get(
    "/dashboard/admin",
    response_model=AdminDashboardResponse,
    tags=["Dashboard"],
)
def admin_dashboard(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.admin:
        raise HTTPException(
            status_code=403,
            detail="Only admins can access this dashboard."
        )

    stats = get_dashboard_stats(db)

    total_users = db.execute(
        select(func.count(User.id))
    ).scalar_one()

    active_users = db.execute(
        select(func.count(User.id))
        .where(User.is_active == True)
    ).scalar_one()

    role_counts = dict(
        db.execute(
            select(
                User.role,
                func.count(User.id)
            )
            .group_by(User.role)
        ).all()
    )

    total_alternatives = db.execute(
        select(func.count(Alternative.id))
    ).scalar_one()

    return AdminDashboardResponse(
        total_users=total_users,
        active_users=active_users,
        employees=role_counts.get(UserRole.employee, 0),
        reviewers=role_counts.get(UserRole.reviewer, 0),
        managers=role_counts.get(UserRole.manager, 0),
        admins=role_counts.get(UserRole.admin, 0),
        total_alternatives=total_alternatives,
        total_decisions=stats["total_decisions"],
        draft_decisions=stats["draft_decisions"],
        under_review_decisions=stats["under_review_decisions"],
        approved_decisions=stats["approved_decisions"],
        rejected_decisions=stats["rejected_decisions"],
        archived_decisions=stats["archived_decisions"],
        recent_decisions=stats["recent_decisions"],
    )

# ============================================================
# Decision Reports
# ============================================================

@app.get(
    "/reports/decision",
    response_model=DecisionReportResponse,
    tags=["Reports"],
)
def get_decision_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in (UserRole.manager, UserRole.admin):
        raise HTTPException(
            status_code=403,
            detail="Only managers and admins can access reports."
        )

    total_decisions = db.execute(
        select(func.count(Decision.id))
    ).scalar_one()

    status_rows = db.execute(
        select(
            Decision.status,
            func.count(Decision.id)
        )
        .group_by(Decision.status)
    ).all()

    by_status = [
        {
            "status": status.value,
            "count": count,
        }
        for status, count in status_rows
    ]

    category_rows = db.execute(
        select(
            Decision.category,
            func.count(Decision.id)
        )
        .where(Decision.category.is_not(None))
        .group_by(Decision.category)
        .order_by(func.count(Decision.id).desc())
    ).all()

    by_category = [
        {
            "category": category,
            "count": count,
        }
        for category, count in category_rows
    ]

    time_rows = db.execute(
        select(
            func.date(Decision.created_at).label("period"),
            func.count(Decision.id)
        )
        .group_by(func.date(Decision.created_at))
        .order_by(func.date(Decision.created_at))
    ).all()

    created_over_time = [
        {
            "period": str(period),
            "count": count,
        }
        for period, count in time_rows
    ]

    recent_decisions = db.execute(
        select(Decision)
        .order_by(Decision.created_at.desc())
        .limit(10)
    ).scalars().all()

    recent_data = [
        {
            "id": decision.id,
            "title": decision.title,
            "status": decision.status.value,
            "category": decision.category,
            "created_at": decision.created_at,
        }
        for decision in recent_decisions
    ]

    return {
        "total_decisions": total_decisions,
        "by_status": by_status,
        "by_category": by_category,
        "created_over_time": created_over_time,
        "recent_decisions": recent_data,
    }

# ============================================================
# Approval Reports
# ============================================================

@app.get(
    "/reports/approvals",
    response_model=ApprovalReportResponse,
    tags=["Reports"],
)
def get_approval_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in (UserRole.manager, UserRole.admin):
        raise HTTPException(
            status_code=403,
            detail="Only managers and admins can access reports."
        )

    total_approvals = db.execute(
        select(func.count(Approval.id))
    ).scalar_one()

    approved = db.execute(
        select(func.count(Approval.id))
        .where(Approval.action == ApprovalAction.approved)
    ).scalar_one()

    rejected = db.execute(
        select(func.count(Approval.id))
        .where(Approval.action == ApprovalAction.rejected)
    ).scalar_one()

    escalated = db.execute(
        select(func.count(Approval.id))
        .where(Approval.action == ApprovalAction.resubmitted)
    ).scalar_one()

    pending = db.execute(
        select(func.count(Decision.id))
        .where(Decision.status == DecisionStatus.under_review)
    ).scalar_one()

    level_rows = db.execute(
        select(
            Approval.stage,
            Approval.action,
            func.count(Approval.id)
        )
        .group_by(Approval.stage, Approval.action)
        .order_by(Approval.stage)
    ).all()

    levels = {}

    for stage, action, count in level_rows:
        if stage not in levels:
            levels[stage] = {
                "level": stage,
                "pending": 0,
                "approved": 0,
                "rejected": 0,
                "escalated": 0,
            }

        if action == ApprovalAction.approved:
            levels[stage]["approved"] = count
        elif action == ApprovalAction.rejected:
            levels[stage]["rejected"] = count
        elif action == ApprovalAction.resubmitted:
            levels[stage]["escalated"] = count

    by_level = list(levels.values())

    return {
        "total_approvals": total_approvals,
        "pending": pending,
        "approved": approved,
        "rejected": rejected,
        "escalated": escalated,
        "average_completion_hours": None,
        "by_level": by_level,
    }

# ============================================================
# Audit Report
# ============================================================

@app.get(
    "/reports/audit",
    response_model=AuditReportResponse,
    tags=["Reports"],
)
def get_audit_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.admin)),
):
    total_events = db.execute(
        select(func.count(AuditLog.id))
    ).scalar_one()

    action_rows = db.execute(
        select(
            AuditLog.action,
            func.count(AuditLog.id)
        )
        .group_by(AuditLog.action)
        .order_by(func.count(AuditLog.id).desc())
    ).all()

    by_action = [
        {
            "action": action,
            "count": count
        }
        for action, count in action_rows
    ]

    actor_rows = db.execute(
        select(
            User.id,
            User.full_name,
            func.count(AuditLog.id)
        )
        .join(AuditLog, AuditLog.user_id == User.id)
        .group_by(User.id, User.full_name)
        .order_by(func.count(AuditLog.id).desc())
    ).all()

    by_actor = [
        {
            "actor_id": actor_id,
            "actor_name": actor_name,
            "count": count
        }
        for actor_id, actor_name, count in actor_rows
    ]

    timeline_rows = db.execute(
        select(
            func.date(AuditLog.created_at),
            func.count(AuditLog.id)
        )
        .group_by(func.date(AuditLog.created_at))
        .order_by(func.date(AuditLog.created_at))
    ).all()

    timeline = [
        {
            "period": str(period),
            "count": count
        }
        for period, count in timeline_rows
    ]

    security_actions = [
        "login",
        "logout",
        "password_change",
        "role_change",
        "user_deleted",
        "user_created"
    ]

    security_rows = db.execute(
        select(AuditLog)
        .where(
            func.lower(AuditLog.action).in_(
                [action.lower() for action in security_actions]
            )
        )
        .order_by(AuditLog.created_at.desc())
        .limit(20)
    ).scalars().all()

    security_events = [
        {
            "id": event.id,
            "action": event.action,
            "actor": (
                {
                    "id": event.user.id,
                    "full_name": event.user.full_name,
                    "email": event.user.email
                }
                if event.user
                else None
            ),
            "created_at": event.created_at
        }
        for event in security_rows
    ]

    recent_rows = db.execute(
        select(AuditLog)
        .order_by(AuditLog.created_at.desc())
        .limit(20)
    ).scalars().all()

    recent_events = [
        {
            "id": event.id,
            "action": event.action,
            "entity_type": event.entity_type,
            "actor": (
                {
                    "id": event.user.id,
                    "full_name": event.user.full_name,
                    "email": event.user.email
                }
                if event.user
                else None
            ),
            "created_at": event.created_at
        }
        for event in recent_rows
    ]

    return AuditReportResponse(
        total_events=total_events,
        by_action=by_action,
        by_actor=by_actor,
        timeline=timeline,
        security_events=security_events,
        recent_events=recent_events,
    )

# ============================================================
# Report Export (PDF)
# ============================================================

def _report_pdf_lines(report_type: str, data: dict) -> list[tuple[str, dict]]:
    """Turn a report's already-computed dict into (text, style) lines for the PDF."""
    lines: list[tuple[str, dict]] = []

    def add(text, size=11, bold=False, gap=0.28):
        lines.append((text, {"size": size, "bold": bold, "gap": gap}))

    if report_type == "decision":
        add(f"Total Decisions: {data['total_decisions']}", bold=True, gap=0.35)
        add("By Status:", bold=True)
        for row in data["by_status"]:
            add(f"  - {row['status'].replace('_', ' ').title()}: {row['count']}")
        add("By Category:", bold=True, gap=0.3)
        for row in data["by_category"]:
            add(f"  - {row['category']}: {row['count']}")
        add("Recent Decisions:", bold=True, gap=0.3)
        for row in data["recent_decisions"]:
            add(f"  - #{row['id']} {row['title']} ({row['status']})")

    elif report_type == "approvals":
        add(f"Total Approvals: {data['total_approvals']}", bold=True, gap=0.35)
        add(f"Pending: {data['pending']}   Approved: {data['approved']}   "
            f"Rejected: {data['rejected']}   Escalated: {data['escalated']}")
        if data.get("average_completion_hours") is not None:
            add(f"Average Completion Time: {data['average_completion_hours']:.1f} hours", gap=0.35)
        add("By Level:", bold=True, gap=0.3)
        for row in data["by_level"]:
            add(f"  - Level {row['level']}: pending {row['pending']}, approved {row['approved']}, "
                f"rejected {row['rejected']}, escalated {row['escalated']}")

    elif report_type == "audit":
        add(f"Total Events: {data['total_events']}", bold=True, gap=0.35)
        add("By Action:", bold=True)
        for row in data["by_action"]:
            add(f"  - {row['action']}: {row['count']}")
        add("By Actor:", bold=True, gap=0.3)
        for row in data["by_actor"]:
            add(f"  - {row['actor_name']}: {row['count']}")
        add("Recent Events:", bold=True, gap=0.3)
        for row in data["recent_events"][:15]:
            actor = row["actor"]["full_name"] if row["actor"] else "Unknown"
            add(f"  - {row['action']} by {actor}")

    else:
        raise HTTPException(status_code=404, detail=f"Unknown report type: {report_type}")

    return lines


REPORT_TITLES = {
    "decision": "Decision Report",
    "approvals": "Approval Report",
    "audit": "Audit Report",
}


def _report_excel_sections(report_type: str, data: dict) -> list[dict]:
    """Turn a report's dict into tabular sections: [{title, headers, rows}]."""
    sections: list[dict] = []

    if report_type == "decision":
        sections.append({
            "title": "Summary",
            "headers": ["Metric", "Value"],
            "rows": [["Total Decisions", data["total_decisions"]]],
        })
        sections.append({
            "title": "By Status",
            "headers": ["Status", "Count"],
            "rows": [[row["status"].replace("_", " ").title(), row["count"]] for row in data["by_status"]],
        })
        sections.append({
            "title": "By Category",
            "headers": ["Category", "Count"],
            "rows": [[row["category"], row["count"]] for row in data["by_category"]],
        })
        sections.append({
            "title": "Recent Decisions",
            "headers": ["ID", "Title", "Status", "Category"],
            "rows": [[row["id"], row["title"], row["status"], row["category"]] for row in data["recent_decisions"]],
        })

    elif report_type == "approvals":
        sections.append({
            "title": "Summary",
            "headers": ["Metric", "Value"],
            "rows": [
                ["Total Approvals", data["total_approvals"]],
                ["Pending", data["pending"]],
                ["Approved", data["approved"]],
                ["Rejected", data["rejected"]],
                ["Escalated", data["escalated"]],
                ["Average Completion (hrs)", data.get("average_completion_hours") if data.get("average_completion_hours") is not None else "N/A"],
            ],
        })
        sections.append({
            "title": "By Level",
            "headers": ["Level", "Pending", "Approved", "Rejected", "Escalated"],
            "rows": [
                [row["level"], row["pending"], row["approved"], row["rejected"], row["escalated"]]
                for row in data["by_level"]
            ],
        })

    elif report_type == "audit":
        sections.append({
            "title": "Summary",
            "headers": ["Metric", "Value"],
            "rows": [["Total Events", data["total_events"]]],
        })
        sections.append({
            "title": "By Action",
            "headers": ["Action", "Count"],
            "rows": [[row["action"], row["count"]] for row in data["by_action"]],
        })
        sections.append({
            "title": "By Actor",
            "headers": ["Actor", "Count"],
            "rows": [[row["actor_name"], row["count"]] for row in data["by_actor"]],
        })
        sections.append({
            "title": "Recent Events",
            "headers": ["Action", "Entity Type", "Actor", "Created At"],
            "rows": [
                [
                    row["action"],
                    row["entity_type"],
                    row["actor"]["full_name"] if row["actor"] else "Unknown",
                    row["created_at"].strftime("%d %b %Y, %I:%M %p") if hasattr(row["created_at"], "strftime") else row["created_at"],
                ]
                for row in data["recent_events"]
            ],
        })

    else:
        raise HTTPException(status_code=404, detail=f"Unknown report type: {report_type}")

    return sections


def _build_report_pdf(report_type: str, data: dict) -> io.BytesIO:
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    y = height - 1 * inch

    def line(text, size=11, bold=False, gap=0.28):
        nonlocal y, p
        if y < 1 * inch:
            p.showPage()
            y = height - 1 * inch
        p.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        p.drawString(1 * inch, y, text)
        y -= gap * inch

    line("Expert Decision Replay Platform", size=10, bold=True)
    line(REPORT_TITLES[report_type], size=16, bold=True, gap=0.4)
    line(f"Generated on: {datetime.utcnow().strftime('%d %b %Y, %I:%M %p')} UTC", size=9, gap=0.4)

    for text, style in _report_pdf_lines(report_type, data):
        line(text, **style)

    p.setFont("Helvetica-Oblique", 8)
    p.drawString(1 * inch, 0.6 * inch, "Generated from Expert Decision Replay Platform")

    p.save()
    buffer.seek(0)
    return buffer


def _build_report_excel(report_type: str, data: dict) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = REPORT_TITLES[report_type][:31]  # sheet names capped at 31 chars

    title_font = Font(bold=True, size=14)
    meta_font = Font(italic=True, size=9, color="666666")
    section_font = Font(bold=True, size=12, color="FFFFFF")
    section_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    row_idx = 1
    ws.cell(row=row_idx, column=1, value=REPORT_TITLES[report_type]).font = title_font
    row_idx += 1
    ws.cell(
        row=row_idx, column=1,
        value=f"Generated on: {datetime.utcnow().strftime('%d %b %Y, %I:%M %p')} UTC",
    ).font = meta_font
    row_idx += 2

    max_cols = 1
    for section in _report_excel_sections(report_type, data):
        max_cols = max(max_cols, len(section["headers"]))

        cell = ws.cell(row=row_idx, column=1, value=section["title"])
        cell.font = section_font
        cell.fill = section_fill
        for c in range(2, len(section["headers"]) + 1):
            ws.cell(row=row_idx, column=c).fill = section_fill
        row_idx += 1

        for c, header in enumerate(section["headers"], start=1):
            cell = ws.cell(row=row_idx, column=c, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row_idx += 1

        if section["rows"]:
            for row_data in section["rows"]:
                for c, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=c, value=value)
                row_idx += 1
        else:
            ws.cell(row=row_idx, column=1, value="No data")
            row_idx += 1

        row_idx += 1  # blank row between sections

    for c in range(1, max_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 24

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@app.get("/reports/{report_type}/export/{export_format}", tags=["Reports"])
def export_report(
    report_type: str,
    export_format: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if report_type not in REPORT_TITLES:
        raise HTTPException(status_code=404, detail=f"Unknown report type: {report_type}")

    if export_format not in ("pdf", "excel"):
        raise HTTPException(status_code=400, detail="Supported export formats: pdf, excel")

    # Re-run the same access checks + data assembly as the on-screen report,
    # then hand the resulting dict to the shared PDF/Excel renderer.
    if report_type == "decision":
        report = get_decision_report(db=db, current_user=current_user)
    elif report_type == "approvals":
        report = get_approval_report(db=db, current_user=current_user)
    elif report_type == "audit":
        report = get_audit_report(db=db, current_user=current_user)

    data = report if isinstance(report, dict) else report.model_dump()

    if export_format == "pdf":
        buffer = _build_report_pdf(report_type, data)
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={report_type}_report.pdf"}
        )
    else:
        buffer = _build_report_excel(report_type, data)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={report_type}_report.xlsx"}
        )